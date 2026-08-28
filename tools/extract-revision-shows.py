"""Extract the set of shows present in each downloaded revision of the sheet.

Layouts changed a lot over four years, so this does not assume fixed columns.
It reads every cell of every sheet and keeps strings that look like a show
name, then the caller diffs consecutive revisions to find additions and
removals.

Written to be over-inclusive here and filtered afterwards: a missed show is
invisible, whereas noise is obvious when the diff is eyeballed.
"""
import json
import os
import re
import sys

from openpyxl import load_workbook

REV_DIR = r"C:/Users/yischoen/.copilot/session-state/7b7cab1a-e3ed-4d93-8833-3f514326952c/files/revisions"

# Rows that are structure, notes or timings rather than shows.
NOISE = re.compile(
    r"^(time|source|duration|when|notes?|todo|ideas?|not done|already doing|removed"
    r"|consider(ing)?|more ideas|daily|weekly|news|sunday|monday|tuesday|wednesday"
    r"|thursday|friday|saturday|shabbat|shabbos|matza|day|min(utes)?|hours?|\d+"
    r"|.*gets home.*|.*fridays are different.*)$",
    re.I,
)
TIME = re.compile(r"^\d{1,2}[:.]\d{2}")
DURATION = re.compile(r"^\d+\s*-\s*\d+$")


def clean(value):
    """A show name from a cell, or None."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or len(s) < 4 or len(s) > 70:
        return None
    # Drop a trailing parenthetical, which is where durations and day patterns
    # live: "Life Kit (15 - 25 minutes, generally Mon + Tue + Thu)".
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()
    s = s.strip(" -–—:•")
    if not s or len(s) < 4:
        return None
    if TIME.match(s) or DURATION.match(s) or NOISE.match(s):
        return None
    # Sentences are notes, not show names.
    if s.count(" ") > 9 or s.endswith("."):
        return None
    if "http" in s.lower() and " " not in s:
        return None
    return s


def norm(s):
    s = s.lower()
    s = re.sub(r"\bthe\b|\bpodcast\b|\bshow\b|\bwith\b|\bby\b|\ba\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


SOURCE_TABS = {"weekly", "daily", "news"}


def _canonical_names():
    """Known show names, longest first, normalised for matching.

    The diff keys a show by whatever text survived cleaning, so the key moved
    whenever the wording around the name moved. Two cases did real damage:
    a cell longer than 70 characters was dropped outright, and the trailing
    parenthetical was only stripped when it ended the cell, which it does not
    when a TODO follows it. So "The Q & A with Rabbi Breitowitz Podcast (40 -
    120 minutes, Tues+Thurs) Change to twice a week" produced no key at all,
    and tidying that note away later looked exactly like adding the show.

    Matching each cell to a known name instead makes the key stable no matter
    what is written around it.
    """
    here = os.path.dirname(os.path.abspath(__file__))
    names = set()
    legacy = os.path.join(here, "sheet-legacy.json")
    if os.path.exists(legacy):
        with open(legacy, encoding="utf8") as fh:
            data = json.load(fh)
        names.update(data.get("notes", {}))
        names.update(data.get("durations", {}))
        names.update(data.get("slots", {}))
    listing = os.path.join(here, "podcast-list.mjs")
    if os.path.exists(listing):
        with open(listing, encoding="utf8") as fh:
            names.update(re.findall(r'name:\s*"([^"]+)"', fh.read()))
    out = {}
    for n in names:
        k = norm(n)
        # Very short keys match far too much: "wired" would swallow every WIRED
        # feed, and the diff would stop distinguishing them.
        if len(k) < 6:
            continue
        # Several sources list the same show, and two spellings that normalise
        # alike must not become two entries: the "points at exactly one show"
        # test below counts them separately and then refuses a good match.
        # Keep the longest spelling, which reads best in the changelog.
        if k not in out or len(n) > len(out[k]):
            out[k] = n
    return sorted(out.items(), key=lambda p: -len(p[0]))


CANON = _canonical_names()

# Renames. The sheet used one name for a show and later another, and without
# this the diff reports the new spelling as a brand new show years after the
# real addition. Kept as an explicit list rather than inferred, because a
# near-identical name is just as often a genuinely different feed: Torat Imecha
# publishes both a Nach Yomi and a Parsha podcast.
ALIASES = {
    # Aleph Beta's parsha podcast, added 2022-12-10 and relabelled 2025-11-27.
    "into verse aleph beta": "into verse parsha",
    "into verse": "into verse parsha",
}


def canonical(value):
    """(key, display name) for a cell naming a known show, else (None, None)."""
    s = str(value)
    s = re.sub(r"https?://\S+", " ", s)
    # Every parenthetical, wherever it sits - not just a trailing one.
    s = re.sub(r"\([^)]*\)", " ", s)
    k = norm(s)
    if len(k) < 6:
        return None, None
    if k in ALIASES:
        k = ALIASES[k]
    for cn, disp in CANON:
        if k == cn or k.startswith(cn + " "):
            return cn, disp
    # A cell may hold an older, shorter form of a name that has since grown:
    # "Into the Verse (Aleph Beta)" became "Into the Verse - A Parsha Podcast",
    # and treating those as two shows invented an addition four years after the
    # real one. Accept the shorter form only when it points at exactly one known
    # show, so genuinely distinct siblings stay distinct - "torat imecha" leads
    # to both the Nach Yomi and the Parsha feeds and is left alone.
    if len(k) >= 8:
        starts = [(cn, disp) for cn, disp in CANON if cn.startswith(k + " ")]
        if len(starts) == 1:
            return starts[0]
    return None, None


def shows_in(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    found = {}
    for ws in wb.worksheets:
        # Only the three tabs that have existed throughout. The catalog and
        # schedule tabs added in 2026 carry hundreds of names and descriptions
        # that would swamp the diff with things that were never scheduled.
        if ws.title.strip().lower() not in SOURCE_TABS:
            continue
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                key, disp = canonical(cell)
                if key:
                    found.setdefault(key, disp)
                    continue
                name = clean(cell)
                if not name:
                    continue
                k = norm(name)
                if len(k) >= 4:
                    found.setdefault(k, name)
    wb.close()
    return found


def main():
    index = json.load(open(os.path.join(REV_DIR, "index.json"), encoding="utf8"))
    out = []
    for rev in index:
        if not os.path.exists(rev["file"]):
            continue
        try:
            found = shows_in(rev["file"])
        except Exception as exc:  # a corrupt export must not lose the rest
            print(f"  {rev['date']} rev {rev['id']}: FAILED {exc}", file=sys.stderr)
            continue
        out.append({**rev, "shows": found})
        print(f"  {rev['date']} rev {str(rev['id']).rjust(5)}  {len(found)} shows")
    with open("revision-shows.json", "w", encoding="utf8") as fh:
        json.dump(out, fh, indent=1, ensure_ascii=False)
    print(f"\nwrote revision-shows.json ({len(out)} revisions)")


if __name__ == "__main__":
    main()
